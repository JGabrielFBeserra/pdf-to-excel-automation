import pdfplumber  # leitor de pdf via table
import pyperclip  # mexer com xlsx
from openpyxl import load_workbook
from copy import copy  # copiar estilos
from tkinter import Tk, filedialog  # abrir caixa de diálogo
import os  # abrir a planilha
import time
from time import sleep
import re


Tk().withdraw()

arquivos_pdf = filedialog.askopenfilenames(
    title="Selecione os PDFs",
    filetypes=[("Arquivos PDF", "*.pdf")]
)

print(arquivos_pdf) # array de caminhos dos pdfs



wb = load_workbook(r"C:\Users\joao.beserra\Documents\sala-tecnica.xlsx")
ws = wb["Planilha1"]

contador = 0
page = 1
# para cada caminho|pdf dentro do array de caminhos
inicio = time.time()
for caminho_pdf in arquivos_pdf:
    # vai abrir cada index|pdf e setar e ext
    with pdfplumber.open(caminho_pdf) as pdf:
        
        #vai ler somente a primeira pagina
        for pagina in pdf.pages:
            print(f"numero da pagina = {page}")
            page += 1
            #extrair o texto da tabela
            tabela = pagina.extract_table()
            print(tabela[0][0])
            if not tabela or not re.match(r"^\d{5}-\d{2}$", tabela[0][0]):
                print("Página ignorada: nenhuma tabela detectada ou formato inválido")
                continue
                        
            
            primeira_linha = tabela[0]
            
            if not any(celula for celula in primeira_linha if celula and '\n' in celula):
                print("Tabela com layout diferente. Página ignorada.")
                continue  
            
             
            count = 0
            dados_extraidos = {}

            #dentro da uma tabela tem uma linha
            for linha in tabela:
                #dentro de uma linha tem uma celula
                for celula in linha:
                    #se a celula n for vazia e tiver \n na celula
                    if celula is not None and '\n' in celula:
                        #ele separa e pega o primeiro pametro como chave e o segundo como valor
                        chave, valor = celula.split('\n', 1)
                        # array de chaves que devem ir pra planilha
                        if chave in ["Data", "Data/Prazo", "Rodovia", "Trecho", "Sentido", "Classificação", "Tipo", "Descrição", "Origem", "Executor" ]:
                            # faz um dicionando adicionando a chave e o vlaor da vez, rodando no for, {'Data': '11/06/2025'}
                            dados_extraidos[chave] = valor

                            

                #count para apenas ler a parte da tabela eu quero
                if count == 10:
                    break
                count += 1

            #adicionando ordem de serviço
            numero_execucao = tabela[0][0]
            dados_extraidos["Ordem_Serviço"] = numero_execucao
            #ordem das chaves
            
            ordem = ["Data", "Origem", "", "Data", "Data", "Data/Prazo", "", "", "", "", "", "Rodovia", "Trecho", "Sentido", "Classificação", "Tipo", "Descrição", "", "Ordem_Serviço", "Executor","", "", "", ""]

            # array com os valores extraidos das chaves
            valores = []
            
            for chave in ordem:
                if chave == "":
                    valores.append("")  # espaço em branco manual
                else:
                    valores.append(dados_extraidos.get(chave, ""))
        
            # todos as formatações que vem estar na planilha de um jeito especifico
            if "retro" in valores[1].lower():
                valores[2] = "RETRO"


            
            valores[11] = valores[11].replace(" ", "-")
            if "artesp".lower() in valores[1].lower():
                valores[1] = "ARTESP"
            else: 
                valores[1] = "ENGENHARIA" 
            
            valores[12:13] = valores[12].split('<>')
            del valores[13]
            valores[12:13] = valores[12].split('+')
            valores[12:13] = valores[12].split('+')
            valores[14] = valores[14].upper()
            if "E: " in valores[15]:
                try:
                    valores[15] = valores[15].split("E: ")[1].strip()
                except IndexError:
                    print("Aviso: 'E: ' encontrado, mas sem conteúdo após. Ignorando valor.")
                    valores[15] = ""        
            
            valores[15] = valores[15].upper()
            valores[16] = valores[16].upper()
            valores[17] = valores[17].split('\n')[1]
            valores[19] = valores[19].split('-')[0] #separa execucao - e pega só a primeira parte
            valores[20] = valores[20].upper()
            
            print(f"{contador} = {valores} \n")
            contador = contador +1
            #formatação de linhas
            linha_formatada = '\t'.join(valores)
            #contagem das linhas, +1 proxima linha, -1 linha de referencia para pegar as formulas
            linha_destino = ws.max_row + 1
            linha_origem = linha_destino - 1
            linha_atual = linha_destino

            #preenchendo as colunas com os valores do array de valores e ao final adiciona os estilos
            for col, valor in enumerate(valores, start=1):
                celula_origem = ws.cell(row=linha_origem, column=col)
                celula_destino = ws.cell(row=linha_destino, column=col)
                

                if col == 8:
                    formula_mes = f'=UPPER(LEFT(TEXT(F{linha_atual},"mmmm"),1)) & MID(TEXT(F{linha_atual},"mmmm"),2,10)'
                    celula_destino.value = formula_mes
                    
                elif col == 9:
                    formula_prazo_edital = f'=F{linha_atual}-D{linha_atual}'
                    celula_destino.value = formula_prazo_edital
                    
                elif col == 10:
                    formula_execucao = f'=F{linha_atual}-G{linha_atual}'
                    celula_destino.value = formula_execucao
                elif col == 13:
                    celula_destino.value = valor
                    celula_destino.value = int(celula_destino.value)

                elif col == 14:
                    celula_destino.value = valor
                    celula_destino.value = int(celula_destino.value)
                elif col == 19:
                    celula_destino.value = "X"
                elif col == 20:
                    celula_destino.value = valor
                    celula_destino.value = int(celula_destino.value)
    
                elif celula_origem.data_type == "f":
                    celula_destino.value = f"={celula_origem.value}"
                else:
                    celula_destino.value = valor

                #adiciona os estilos das celulas destino para celula origem
                celula_destino.font = copy(celula_origem.font)
                celula_destino.fill = copy(celula_origem.fill)
                celula_destino.border = copy(celula_origem.border)
                celula_destino.alignment = copy(celula_origem.alignment)
                celula_destino.number_format = celula_origem.number_format

# Salva e abre a planilha
fim = time.time()
tempo_total = fim - inicio
print(f"Tempo total de execução: {tempo_total:.2f} segundos")
print("Foram planilhadas", contador, "ordens de serviço!")


while True:
    try:
        wb.save(r"C:\Users\joao.beserra\Documents\sala-tecnica.xlsx")
        print("✅ Arquivo salvo com sucesso! Abrindo a Planilha")
        os.startfile(r"C:\Users\joao.beserra\Documents\sala-tecnica.xlsx")
        break  
    except PermissionError:
        print("\n❌ Não foi possível salvar o arquivo 'sala-tecnica.xlsx'.")
        print("🧐 Verifique se ele está aberto em outro programa e feche para tentar novamente.")
        input("🔁 Pressione Enter para tentar salvar de novo...")
        sleep(1)  
        
